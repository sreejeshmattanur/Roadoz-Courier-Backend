import csv
import io
from datetime import date, datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas

# Column configurations for all report types
REPORT_CONFIGS = {
    "Daily Booking Report": {
        "headers": ["Booking No", "Sender", "Receiver", "Destination", "Weight (kg)", "Amount", "Status"],
        "keys": ["booking_no", "sender", "receiver", "destination", "weight", "amount", "status"]
    },
    "Customer Wise Booking Report": {
        "headers": ["Customer", "Bookings", "Revenue", "Pending Amount"],
        "keys": ["customer", "bookings", "revenue", "pending_amount"]
    },
    "Service Type Report": {
        "headers": ["Service Type", "Total Bookings", "Revenue"],
        "keys": ["service_type", "total_bookings", "revenue"]
    },
    "Delivery Status Report": {
        "headers": ["AWB No", "Receiver", "Delivery Date", "Status", "POD"],
        "keys": ["awb_no", "receiver", "delivery_date", "status", "pod"]
    },
    "Pending Delivery Report": {
        "headers": ["AWB No", "Destination", "Pending Days", "Reason"],
        "keys": ["awb_no", "destination", "pending_days", "reason"]
    },
    "COD Pending Report": {
        "headers": ["Booking No", "Merchant", "Pending Amount", "Status"],
        "keys": ["booking_no", "merchant", "pending_amount", "status"]
    },
    "GST Sales Report": {
        "headers": ["Invoice No", "Taxable Amount", "CGST", "SGST", "IGST", "Total"],
        "keys": ["invoice_no", "taxable_amount", "cgst", "sgst", "igst", "total"]
    },
    "Franchise Settlement Report": {
        "headers": ["Franchise ID", "Franchise", "Shipments", "Revenue", "HO Share", "Franchise Share", "Net Payable"],
        "keys": ["franchise_id", "franchise", "shipments", "revenue", "ho_share", "franchise_share", "net_payable"]
    },
    "Monthly Revenue Analysis": {
        "headers": ["Month", "Revenue", "Expenses", "Profit", "Growth %"],
        "keys": ["month", "revenue", "expenses", "profit", "growth_percent"]
    },
    "Top Customer Report": {
        "headers": ["Customer", "Shipments", "Revenue", "Outstanding"],
        "keys": ["customer", "bookings", "revenue", "pending_amount"]
    },
    "Delivery Efficiency Report": {
        "headers": ["Branch", "Assigned", "Delivered", "Pending", "Efficiency %"],
        "keys": ["branch", "assigned", "delivered", "pending", "efficiency_percent"]
    },
    "Day Close Report": {
        "headers": ["Opening Balance", "Collection", "Expenses", "Closing Balance"],
        "keys": ["opening_balance", "collection", "expenses", "closing_balance"]
    },
    "Branch Activity Report": {
        "headers": ["Branch", "Bookings", "Deliveries", "Collections", "Pending"],
        "keys": ["branch", "bookings", "deliveries", "collections", "pending"]
    },
    "User Activity Report": {
        "headers": ["User", "Login Time", "Logout Time", "Transactions"],
        "keys": ["user", "login_time", "logout_time", "transactions"]
    },
    "Returned Shipment Report": {
        "headers": ["AWB No", "Return Reason", "Status"],
        "keys": ["awb_no", "return_reason", "status"]
    },
    "Collection Summary Report": {
        "headers": ["Receipt No", "Customer", "Amount", "Payment Mode"],
        "keys": ["receipt_no", "customer", "amount", "payment_mode"]
    },
    "Outstanding Collection Report": {
        "headers": ["Customer", "Invoice", "Balance", "Due Date"],
        "keys": ["customer", "invoice", "balance", "due_date"]
    },
    "Daily Collection Report": {
        "headers": ["Date", "Cash", "UPI", "Bank Transfer", "Total"],
        "keys": ["date", "cash", "upi", "bank_transfer", "total"]
    },
    "COD Settlement Report": {
        "headers": ["Merchant", "COD Amount", "Commission", "Net Payable"],
        "keys": ["merchant", "cod_amount", "commission", "net_payable"]
    },
    "COD Commission Report": {
        "headers": ["Merchant", "Commission %", "Amount"],
        "keys": ["merchant", "commission_percent", "amount"]
    },
    "Cash Book Report": {
        "headers": ["Date", "Voucher No", "Debit", "Credit", "Balance"],
        "keys": ["date", "voucher_no", "debit", "credit", "balance"]
    },
    "Expense Report": {
        "headers": ["Expense Head", "Amount", "Approved By"],
        "keys": ["expense_head", "amount", "approved_by"]
    },
    "Profit & Loss Report": {
        "headers": ["Revenue", "Expenses", "Net Profit"],
        "keys": ["revenue", "expenses", "net_profit"]
    },
    "HSN Summary Report": {
        "headers": ["HSN Code", "Taxable Amount", "GST Amount"],
        "keys": ["hsn_code", "taxable_amount", "gst_amount"]
    },
    "GST Collection Summary": {
        "headers": ["Month", "Collected GST", "Paid GST", "Balance"],
        "keys": ["month", "collected_gst", "paid_gst", "balance"]
    },
    "Franchise Outstanding Report": {
        "headers": ["Franchise", "Pending Amount", "Due Date", "Status"],
        "keys": ["franchise", "pending_amount", "due_date", "status"]
    },
    "Franchise Collection Report": {
        "headers": ["Franchise", "Cash Collection", "Bank Deposit", "Closing Balance"],
        "keys": ["franchise", "cash_collection", "bank_deposit", "closing_balance"]
    },
    "Franchise Profitability Report": {
        "headers": ["Franchise", "Revenue", "Expenses", "Profit"],
        "keys": ["franchise", "revenue", "expenses", "profit"]
    },
    "Area Wise Business Report": {
        "headers": ["Area", "Shipments", "Revenue", "Pending Deliveries"],
        "keys": ["area", "shipments", "revenue", "pending_deliveries"]
    },
    "Performance Dashboard Report": {
        "headers": ["Parameter", "Current Month", "Previous Month", "Growth %"],
        "keys": ["parameter", "current_month", "previous_month", "growth_percent"]
    }
}


def _resolve_config(title: str, items: list[dict]) -> tuple[list[str], list[str]]:
    config = REPORT_CONFIGS.get(title)
    if config:
        return config["headers"], config["keys"]
    if items:
        keys = list(items[0].keys())
        headers = [k.replace("_", " ").title() for k in keys]
        return headers, keys
    return ["No Data"], ["no_data"]


# ── CSV EXPORTER ──────────────────────────────────────────────────────────────
def export_to_csv(report_data: dict) -> bytes:
    title = report_data.get("report", "Report")
    items = report_data.get("items", [])
    
    headers, keys = _resolve_config(title, items)
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Metadata
    writer.writerow([f"ROADOZ COURIER & CARGO"])
    writer.writerow([title.upper()])
    if "date" in report_data:
        writer.writerow([f"Date: {report_data['date']}"])
    if "year" in report_data:
        writer.writerow([f"Year: {report_data['year']}"])
    writer.writerow([])
    
    # Headers
    writer.writerow(headers)
    
    # Data Rows
    for item in items:
        writer.writerow([item.get(k, "") for k in keys])
        
    return output.getvalue().encode("utf-8")


# ── EXCEL EXPORTER ────────────────────────────────────────────────────────────
def export_to_excel(report_data: dict) -> bytes:
    title = report_data.get("report", "Report")
    items = report_data.get("items", [])
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    
    # Enable gridlines explicitly
    ws.views.sheetView[0].showGridLines = True
    
    # Premium Typography and Palette (Corporate Indigo)
    font_title = Font(name="Calibri", size=16, bold=True, color="1F4E79")
    font_subtitle = Font(name="Calibri", size=10, bold=True, color="595959")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Calibri", size=11)
    font_meta = Font(name="Calibri", size=10, italic=True, color="595959")
    
    fill_header = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    fill_zebra = PatternFill(start_color="F2F7FA", end_color="F2F7FA", fill_type="solid")
    
    border_thin = Side(border_style="thin", color="D9D9D9")
    border_cell = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    
    # Header block
    ws.cell(row=2, column=2, value="ROADOZ COURIER & CARGO").font = font_subtitle
    ws.cell(row=3, column=2, value=title.upper()).font = font_title
    
    meta_row = 4
    if "date" in report_data:
        ws.cell(row=meta_row, column=2, value=f"Report Date: {report_data['date']}").font = font_meta
        meta_row += 1
    if "year" in report_data:
        ws.cell(row=meta_row, column=2, value=f"Year: {report_data['year']}").font = font_meta
        meta_row += 1
        
    start_row = meta_row + 1
    
    headers, keys = _resolve_config(title, items)
    
    # Headers rendering
    for col_idx, header in enumerate(headers, start=2):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_cell
        
    # Data rendering
    current_row = start_row + 1
    for idx, item in enumerate(items):
        for col_idx, key in enumerate(keys, start=2):
            val = item.get(key, "")
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = font_data
            cell.border = border_cell
            
            # Alignments & custom formatting
            if isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal="right")
                if "percent" in key or "growth" in key:
                    cell.number_format = "0.00'%'"
                elif isinstance(val, float) or "amount" in key or "revenue" in key or "charge" in key or "payable" in key or "balance" in key or "expenses" in key or "profit" in key or "gst" in key or "collection" in key or "deposit" in key:
                    cell.number_format = "#,##0.00"
            elif "date" in key or "time" in key:
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="left")
                
            # Zebra pattern
            if idx % 2 == 1:
                cell.fill = fill_zebra
                
        current_row += 1
        
    # Auto-adjust column widths
    for col in ws.columns:
        if col[0].column < 2 or col[0].column > len(headers) + 1:
            continue
        max_len = 0
        for cell in col:
            if cell.row >= start_row and cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# ── PDF EXPORTER (NumberedCanvas Pattern) ──────────────────────────────────────
class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_decorations(num_pages)
            super().showPage()
        super().save()

    def draw_page_decorations(self, page_count):
        self.saveState()
        self.setFont("Helvetica-Bold", 8)
        self.setFillColor(colors.HexColor("#595959"))
        
        # Header banner line
        self.setLineWidth(0.75)
        self.setStrokeColor(colors.HexColor("#1F4E79"))
        self.line(36, 800, 559, 800)
        self.drawString(36, 805, "ROADOZ COURIER & CARGO SYSTEMS")
        self.drawRightString(559, 805, "OFFICIAL BUSINESS REPORT")
        
        # Footer
        self.line(36, 45, 559, 45)
        self.setFont("Helvetica", 8)
        self.drawString(36, 32, "Confidential Document - Generated Automatically")
        page_text = f"Page {self._pageNumber} of {page_count}"
        self.drawRightString(559, 32, page_text)
        self.restoreState()


def export_to_pdf(report_data: dict) -> bytes:
    title = report_data.get("report", "Report")
    items = report_data.get("items", [])
    
    # 0.5 inch margins = 36 pt. Available width on A4 (595 x 842) is 523 pt.
    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        pdf_buffer,
        pagesize=A4,
        leftMargin=36,
        rightMargin=36,
        topMargin=54,
        bottomMargin=54
    )
    
    styles = getSampleStyleSheet()
    
    # Unique style configurations
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=18,
        textColor=colors.HexColor("#1F4E79"),
        spaceAfter=4
    )
    subtitle_style = ParagraphStyle(
        "ReportSubTitle",
        parent=styles["Normal"],
        fontName="Helvetica-Oblique",
        fontSize=10,
        textColor=colors.HexColor("#595959"),
        spaceAfter=15
    )
    th_style = ParagraphStyle(
        "TableHeader",
        fontName="Helvetica-Bold",
        fontSize=9,
        textColor=colors.white,
        alignment=1 # Center aligned
    )
    td_style = ParagraphStyle(
        "TableCell",
        fontName="Helvetica",
        fontSize=8.5,
        textColor=colors.HexColor("#333333")
    )
    
    story = []
    
    # Title & Metadata
    story.append(Paragraph(title.upper(), title_style))
    meta_desc = f"Generated on {datetime.now().strftime('%d-%m-%Y %H:%M')}"
    if "date" in report_data:
        meta_desc += f"  |  Report Period: {report_data['date']}"
    if "year" in report_data:
        meta_desc += f"  |  Year: {report_data['year']}"
    story.append(Paragraph(meta_desc, subtitle_style))
    
    headers, keys = _resolve_config(title, items)
    
    # Assemble Table Data
    table_data = []
    # Add Header Row
    table_data.append([Paragraph(h, th_style) for h in headers])
    
    # Add Data Rows
    for item in items:
        row = []
        for key in keys:
            val = item.get(key, "")
            # Formattings
            if isinstance(val, float) or "amount" in key or "revenue" in key or "charge" in key or "payable" in key or "balance" in key or "expenses" in key or "profit" in key or "gst" in key or "collection" in key or "deposit" in key:
                val_str = f"{float(val or 0):,.2f}"
            elif "growth" in key or "percent" in key:
                val_str = f"{float(val or 0):.2f}%"
            else:
                val_str = str("" if val is None else val)
                
            row.append(Paragraph(val_str, td_style))
        table_data.append(row)
        
    # Beautiful Table Styling
    col_width = 523 / len(headers)
    t = Table(table_data, colWidths=[col_width] * len(headers))
    t_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
    ])
    
    # Add Zebra striping
    for i in range(1, len(table_data)):
        if i % 2 == 0:
            t_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor("#F8FAFC"))
            
    t.setStyle(t_style)
    story.append(t)
    
    # Build PDF with dynamic page canvas
    doc.build(story, canvasmaker=NumberedCanvas)
    pdf_bytes = pdf_buffer.getvalue()
    pdf_buffer.close()
    return pdf_bytes
